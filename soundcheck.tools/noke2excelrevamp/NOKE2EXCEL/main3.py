import concurrent.futures
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

# Windows console defaults to a non-UTF-8 codepage, so the arrows/dashes below
# raise UnicodeEncodeError there (it's a ValueError subclass, so login() was
# silently swallowing it and misreporting successful logins as "bad JSON").
sys.stdout.reconfigure(encoding="utf-8")

#theheadbitchinchargeiscodinghoe

# ─── Config ───────────────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")

EMAIL = os.environ["NOKE_EMAIL"]
PASSWORD = os.environ["NOKE_PASSWORD"]
BASE_URL = "https://router.smartentry.noke.com"
SHEET_PROTECTION_PASSWORD = os.environ["SHEET_PROTECTION_PASSWORD"]
LOGIN_PATH = "/login/"
UNITS_PATH = "/site/units/"
PROFILE_PATH = "/lock/profile/"
UNIT_FILTERS = [
   "available",
   "inuse",
   "pending",
   "auction-smartentry",
   "transfer-smartentry"
]

# Where the spreadsheets are written on this machine (server: ~/noke2excel/output).
# rclone uploads everything in this folder to the shared OneDrive folder.
OUTPUT_DIR = Path(__file__).parent / "spreads"

# How often to refresh, in seconds (300 = every 5 minutes)
REFRESH_SECONDS = 300
# ────────────────────────────────────────────────────────────────────────────

def login() -> str | None:
   url = BASE_URL + LOGIN_PATH
   payload = {
       "companyUUID": "-1",
       "email": EMAIL,
       "password": PASSWORD
   }
   try:
       r = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=12)
       r.raise_for_status()
       data = r.json()
       token = data.get("token")
       if token:
           print("Login successful → new token received")
           return token
       print("Login response missing 'token' field")
       return None
   except requests.RequestException as e:
       print(f"Login failed: {e}")
       if hasattr(e, 'response') and e.response is not None:
           print("Error body:", e.response.text)
       return None
   except ValueError as e:
       print(f"Bad JSON from login: {e}")
       return None

# ─── Fetch shared users for a specific unit ───────────────────────────────
def get_shared_users(token: str, unit_uuid: str) -> list:
   url = BASE_URL + PROFILE_PATH
   headers = {
       "Accept": "application/json, text/plain, */*",
       "Authorization": f"Bearer {token}",
       "Content-Type": "application/json"
   }
   payload = {"unitUUID": unit_uuid}
   try:
       r = requests.post(url, json=payload, headers=headers, timeout=10)
       r.raise_for_status()
       resp = r.json()
       if "data" in resp and "sharedUsers" in resp["data"]:
           shared = []
           for u in resp["data"]["sharedUsers"]:
               shared.append({
                   "first_name": u.get("firstName", ""),
                   "last_name":  u.get("lastName", ""),
                   "email":      u.get("email", ""),
                   "phone":      u.get("phone", ""),
               })
           return shared
       return []
   except Exception as e:
       print(f"Failed to fetch shared users for unit {unit_uuid}: {e}")
       return []

# ─── Parse units into clean list ──────────────────────────────────────────
def extract_unit_info(data, token: str, max_workers: int = 10):
    """
    Extracts rentable unit info in parallel.
    Returns a list of dicts, skipping non‐rentable units.
    """
    if "data" not in data or "units" not in data["data"]:
        print("No 'data.units' found in response")
        return []

    units = data["data"]["units"]

    def process_unit(unit):
        if unit.get("accessType") != "rentable":
            return None

        # User info
        user = unit.get("User", {})

        # Gatecode: first accessCode if present
        gatecode = ""
        access_codes = unit.get("accessCodes") or []
        if isinstance(access_codes, list) and access_codes:
            gatecode = access_codes[0].get("accessCode", "")

        # Shared users: only fetch if shared == True
        shared_users = []
        if unit.get("shared", False):
            unit_uuid = unit.get("uuid")
            if unit_uuid:
                shared_users = get_shared_users(token, unit_uuid)

        return {
            "unit_id":      unit.get("externalId", ""),
            "unit_name":    unit.get("name", ""),
            "rental_state": unit.get("rentalState", ""),
            "first_name":   user.get("firstName", ""),
            "last_name":    user.get("lastName", ""),
            "email":        user.get("email", ""),
            "phone":        user.get("phone", ""),
            "gatecode":     gatecode,
            "shared_users": shared_users,

        }

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = executor.map(process_unit, units)

    # Filter out None (non‐rentable units) and return list
    return [entry for entry in results if entry is not None]

# ─── Excel helpers ─────────────────────────────────────────────────────────
def write_df_preserve_formatting(path: str, df: pd.DataFrame, sheet_name: str = "Sheet1"):
    """
    Write a DataFrame into an existing workbook while preserving
    existing cell formatting as much as possible.

    If the workbook or sheet does not exist, it falls back to a
    normal pandas to_excel call which creates the file.
    """
    file_path = Path(path)

    # If the file doesn't exist yet, just create it normally then re-open to size columns
    if not file_path.exists():
        df.to_excel(path, index=False)
        wb = load_workbook(filename=path)
        ws = wb.active
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 4
        wb.save(path)
        return

    try:
        wb = load_workbook(filename=path)
    except Exception:
        # If we cannot open it with openpyxl, fall back to normal write
        df.to_excel(path, index=False)
        return

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Assume first row is headers with desired formatting; keep it and
    # clear everything below while leaving styles intact for row 1.
    max_row = ws.max_row
    if max_row > 1:
        for row in range(2, max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None

    # Write headers (row 1) – keep existing formatting where possible
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name

    # Write data starting from row 2
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-size columns to fit content
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 4

    # Protect the sheet so only the program can write; openpyxl bypasses this
    ws.protection.sheet = True
    ws.protection.password = SHEET_PROTECTION_PASSWORD
    ws.protection.enable()

    wb.save(path)


# ─── Main flow ─────────────────────────────────────────────────────────────
def run_once():
    print(f"\n=== Refresh run at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    print("Logging in fresh to get new token...")
    token = login()
    if not token:
        print("Login failed — check credentials/network. Skipping this run.")
        return

    url = BASE_URL + UNITS_PATH
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    payload = {"filter": UNIT_FILTERS}
    try:
        r = requests.post(url, json=payload, headers=headers, timeout=10)
        r.raise_for_status()
        response_data = r.json()
        units_list = extract_unit_info(response_data, token)
        if not units_list:
            print("No rentable units found.")
            return

        print(f"\nFound {len(units_list)} rentable units.")

        # Keep `shared_users` as a comma-separated names list (like before),
        # and add `shared_user_phones` as a comma-separated phone list
        # in the exact order returned by the API.
        for u in units_list:
            names = []
            phones = []
            for su in u.get("shared_users", []):
                fn = su.get("first_name", "").strip()
                ln = su.get("last_name", "").strip()
                if fn or ln:
                    names.append(f"{fn} {ln}".strip())

                phone = su.get("phone", "").strip()
                if phone:
                    phones.append(phone)

            u["shared_users"] = ", ".join(names)
            u["shared_user_phones"] = ", ".join(phones)

        print("\nWriting to Excel...")
        df = pd.DataFrame(units_list)

        # Use current month/year in filenames
        today = datetime.today()
        month_year = today.strftime("%B_%Y")  # e.g. "March_2026"

        # Make sure the output folder exists, then always overwrite the
        # main/live spreadsheet for this month, keeping the workbook formatting.
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        live_path = str(OUTPUT_DIR / f"Current_{month_year}.xlsx")
        print(f"Saving live file to: {live_path}")
        write_df_preserve_formatting(live_path, df)

        # On the 19th of each month, also save a crystallized month/year copy
        if today.day == 19:
            archive_path = str(OUTPUT_DIR / f"{month_year}.xlsx")
            print(f"Also saving monthly snapshot to: {archive_path}")
            df.to_excel(archive_path, index=False)

    except requests.RequestException as e:
        print(f"Units request failed: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print("Status:", e.response.status_code)
            print("Body:", e.response.text)
    except ValueError:
        print("Response was not valid JSON")


if __name__ == "__main__":
    while True:
        try:
            run_once()
        except Exception as e:
            print(f"Unexpected error in run: {e}")
        print(f"Sleeping for {REFRESH_SECONDS} seconds...\n")
        time.sleep(REFRESH_SECONDS)  # 5 minutes