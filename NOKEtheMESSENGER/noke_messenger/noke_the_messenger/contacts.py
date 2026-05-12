import openpyxl
import os

EXCEL_PATH = os.getenv("EXCEL_PATH", "contacts.xlsx")

def load_contacts():
    """Load all contacts from the Excel sheet."""
    wb = openpyxl.load_workbook(r"C:\Users\sound\OneDrive\Noke2Excel\Current_March_2026.xlsx")
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    contacts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        # Skip empty rows or unavailable units
        if not row_data.get("unit_id"):
            continue

        # Build primary contact
        primary = {
            "unit_id": row_data.get("unit_id", ""),
            "unit_name": row_data.get("unit_name", ""),
            "rental_state": row_data.get("rental_state", ""),
            "name": f"{row_data.get('first_name', '')} {row_data.get('last_name', '')}".strip(),
            "first_name": row_data.get("first_name", ""),
            "last_name": row_data.get("last_name", ""),
            "email": row_data.get("email", ""),
            "phone": format_phone(row_data.get("phone", "")),
            "contact_type": "primary",
            "opt_out": row_data.get("sms_opt_out", False)
        }

        if primary["name"] and not primary["opt_out"]:
            contacts.append(primary)

        # Build secondary/shared contacts
        shared_names = parse_csv(row_data.get("shared_users", ""))
        shared_phones = parse_csv(row_data.get("shared_phones", ""))
        for i, name in enumerate(shared_names):
            secondary = {
                "unit_id": row_data.get("unit_id", ""),
                "unit_name": row_data.get("unit_name", ""),
                "rental_state": row_data.get("rental_state", ""),
                "name": name.strip(),
                "first_name": name.strip().split(" ")[0] if name else "",
                "last_name": "",
                "phone": format_phone(shared_phones[i]) if i < len(shared_phones) else "",
                "contact_type": "secondary",
                "opt_out": False
            }
            if secondary["name"]:
                contacts.append(secondary)

    return contacts


def get_group(contacts, group="all", camp=None):
    """Filter contacts by group type and optionally by camp."""
    filtered = contacts

    # Filter by camp if specified
    if camp:
        filtered = [c for c in filtered if c["unit_name"] == camp]

    # Filter by contact type
    if group == "primary":
        filtered = [c for c in filtered if c["contact_type"] == "primary"]
    elif group == "secondary":
        filtered = [c for c in filtered if c["contact_type"] == "secondary"]
    # "all" returns everyone

    return filtered


def format_phone(phone):
    """Normalize phone number to E.164 format for Twilio."""
    if not phone:
        return ""
    digits = str(phone).replace("-", "").replace(" ", "").replace("(", "").replace(")", "").replace(".", "")
    # Remove any decimal point (Excel sometimes stores as float)
    digits = digits.split(".")[0]
    if len(digits) == 10:
        return f"+1{digits}"
    elif len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return f"+{digits}"


def parse_csv(value):
    """Parse comma-separated string into list."""
    if not value:
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]
